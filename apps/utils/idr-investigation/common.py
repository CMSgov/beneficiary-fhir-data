import csv
from abc import ABC, abstractmethod
from collections.abc import Sequence
from dataclasses import asdict, dataclass, field, fields
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.cell import Cell


@dataclass
class CsvWriteable(ABC):
    @staticmethod
    @abstractmethod
    def get_header_field_mapping() -> dict[str, str]: ...


@dataclass
class BfdV2DataFieldRow(CsvWriteable):
    column_name: str
    field_name: str
    description: str
    ccw_mapping: str
    cclf_mapping: str

    @staticmethod
    def get_header_field_mapping() -> dict[str, str]:
        return {
            "column_name": "BFD Field Database Column Name",
            "field_name": "BFD Field Name",
            "description": "BFD Field Description",
            "ccw_mapping": "BFD Field CCW Mapping",
            "cclf_mapping": "BFD Field CCLF Mapping",
        }


@dataclass
class RdaDataFieldRow(CsvWriteable):
    source_system: str
    api_field_label: str
    field_name: str
    source_copybook_field: str
    source_system_def: str

    @staticmethod
    def get_header_field_mapping() -> dict[str, str]:
        return {
            "source_system": "RDA Source System",
            "api_field_label": "RDA API Field Label",
            "field_name": "RDA Field Name",
            "source_copybook_field": "RDA Source Copybook Field Label",
            "source_system_def": "RDA Source System Definition",
        }


@dataclass
class IdrDataFieldRow(CsvWriteable):
    column_name: str
    entity_name: str
    col_data_source_name: str
    attribute_name: str
    attribute_def: str
    table_name: str
    view_name: str = field(default="Unknown")

    def __post_init__(self: "IdrDataFieldRow") -> None:
        if self.table_name and not any(x in self.table_name for x in ["CLM", "CLM_LINE"]):
            self.view_name = f"V2_MDCR_{self.table_name}"

    @staticmethod
    def get_header_field_mapping() -> dict[str, str]:
        return {
            "column_name": "IDR Column Name",
            "entity_name": "IDR Entity Name",
            "col_data_source_name": "IDR Column Data Source Name",
            "attribute_name": "IDR Attribute Name",
            "attribute_def": "IDR Attribute Definition",
            "table_name": "IDR Table Name",
            "view_name": "Possible IDR View Name",
        }


@dataclass
class BfdRdaColumnDataFieldRow(CsvWriteable):
    column_name: str

    @staticmethod
    def get_header_field_mapping() -> dict[str, str]:
        return {"column_name": "BFD RDA Schema Column Name"}


def clean_str(str_val: str) -> str:
    oneline_str_val = str_val.replace("\n", " ").replace("\\n", " ").strip()
    return "".join(e for e in oneline_str_val if e.isalnum() or e == " ").lower()


def get_col_idx_from_header_row(sheet_header_row: tuple[Cell, ...], header_label: str) -> int:
    try:
        return next(
            x.col_idx - 1
            for x in sheet_header_row
            if str(x.value).lower().strip() == header_label.lower().strip()
        )
    except StopIteration as err:
        raise RuntimeError("Invalid header label to search for") from err


def cell_str_value_at_col(row: tuple[Cell, ...], col_idx: int) -> str:
    return (
        str(row[col_idx].value or "")
        .encode("unicode_escape")
        .decode()
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\n", "\\n")
        .strip()
    )


def write_to_csv(
    filename: str,
    list_to_write: Sequence[CsvWriteable],
) -> None:
    print(f"Writing {len(list_to_write)} row(s) to {filename}...")
    with Path(filename).open("w") as csv_file:
        wr = csv.writer(csv_file, delimiter=",", dialect="unix")
        wr.writerow(
            header_mapping[x.name]
            for x in fields(list_to_write[0])
            if (header_mapping := list_to_write[0].get_header_field_mapping())
        )
        wr.writerows(
            (
                y if not isinstance(y, list) else ", ".join(cast(list[Any], y))
                for y in asdict(x).values()
            )
            for x in list_to_write
        )
    print(f"Successfully wrote {len(list_to_write)} row(s) to {filename}")


def extract_bfd_rows_from_path(workbook_path: str) -> list[BfdV2DataFieldRow]:
    bfd_wb = load_workbook(workbook_path)
    bfd_sheet = bfd_wb["V2"]
    bfd_header_row = next(bfd_sheet.iter_rows())  # First row is header row
    bfd_field_name_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="Name"
    )
    bfd_description_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="Description"
    )
    bfd_ccw_mapping_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="CCW Mapping"
    )
    bfd_cclf_mapping_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="CCLF Mapping"
    )
    bfd_column_name_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="BFD Column Name"
    )
    return [
        BfdV2DataFieldRow(
            column_name=column_name,
            field_name=cell_str_value_at_col(row=row, col_idx=bfd_field_name_column),
            description=cell_str_value_at_col(row=row, col_idx=bfd_description_column),
            ccw_mapping=cell_str_value_at_col(row=row, col_idx=bfd_ccw_mapping_column),
            cclf_mapping=cell_str_value_at_col(row=row, col_idx=bfd_cclf_mapping_column),
        )
        for row in bfd_sheet.iter_rows(min_row=2)
        # There are a few meta fields that aren't useful and have no map in the DB. Exclude them.
        if (column_name := cell_str_value_at_col(row=row, col_idx=bfd_column_name_column))
    ]


def extract_rda_rows_from_path(workbook_path: str) -> list[RdaDataFieldRow]:
    rda_wb = load_workbook(workbook_path)
    rda_sheet = rda_wb["B) Data Dictionary"]
    rda_header_row = next(rda_sheet.iter_rows())  # First row is header row
    rda_source_system_column = get_col_idx_from_header_row(
        sheet_header_row=rda_header_row, header_label="source system"
    )
    rda_field_name_column = get_col_idx_from_header_row(
        sheet_header_row=rda_header_row, header_label="field name"
    )
    rda_api_field_label_column = get_col_idx_from_header_row(
        sheet_header_row=rda_header_row, header_label="rda api field label"
    )
    rda_copybook_field_column = get_col_idx_from_header_row(
        sheet_header_row=rda_header_row, header_label="source copybook field label"
    )
    rda_source_system_def_column = get_col_idx_from_header_row(
        sheet_header_row=rda_header_row, header_label="source system definition"
    )
    return [
        RdaDataFieldRow(
            source_system=source_system,
            api_field_label=cell_str_value_at_col(row=row, col_idx=rda_api_field_label_column),
            field_name=cell_str_value_at_col(row=row, col_idx=rda_field_name_column),
            source_copybook_field=cell_str_value_at_col(row=row, col_idx=rda_copybook_field_column),
            source_system_def=cell_str_value_at_col(row=row, col_idx=rda_source_system_def_column),
        )
        for row in rda_sheet.iter_rows(min_row=2)
        # RDA DD has one field without a source system which is the Claim Key. We don't need it.
        if (source_system := cell_str_value_at_col(row=row, col_idx=rda_source_system_column))
    ]


def extract_idr_rows_from_path(workbook_path: str) -> list[IdrDataFieldRow]:
    idr_wb = load_workbook(workbook_path)
    idr_sheet = idr_wb["Column_Defs"]
    idr_header_row = next(idr_sheet.iter_rows(min_row=4))  # header is 4th row
    idr_col_data_source_name_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Column Data Source Name"
    )
    idr_column_name_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Column Name"
    )
    idr_entity_name_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Entity Name"
    )
    idr_attribute_name_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Attribute Name"
    )
    idr_attribute_def_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Attribute Definition"
    )
    idr_table_name_column = get_col_idx_from_header_row(
        sheet_header_row=idr_header_row, header_label="Table Name"
    )
    return [
        IdrDataFieldRow(
            column_name=column_name,
            entity_name=cell_str_value_at_col(row=row, col_idx=idr_entity_name_column),
            col_data_source_name=cell_str_value_at_col(
                row=row, col_idx=idr_col_data_source_name_column
            ),
            attribute_name=cell_str_value_at_col(row=row, col_idx=idr_attribute_name_column),
            attribute_def=cell_str_value_at_col(row=row, col_idx=idr_attribute_def_column),
            table_name=cell_str_value_at_col(row=row, col_idx=idr_table_name_column),
        )
        for row in idr_sheet.iter_rows(min_row=5)
        # IDR DD has a few empty rows at the end. We can get rid of them
        if (column_name := cell_str_value_at_col(row=row, col_idx=idr_column_name_column))
    ]


def extract_bfd_rda_columns_from_path(workbook_path: str) -> list[BfdRdaColumnDataFieldRow]:
    bfd_wb = load_workbook(workbook_path)
    bfd_sheet = bfd_wb["Sheet 1"]
    bfd_header_row = next(bfd_sheet.iter_rows(min_row=2))
    bfd_rda_column_name_column = get_col_idx_from_header_row(
        sheet_header_row=bfd_header_row, header_label="column_name"
    )
    return [
        BfdRdaColumnDataFieldRow(
            column_name=cell_str_value_at_col(row=row, col_idx=bfd_rda_column_name_column)
        )
        for row in bfd_sheet.iter_rows(min_row=3)
    ]
