from openpyxl import load_workbook
from openpyxl.styles import Alignment
from thefuzz import fuzz
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

header_font_style = Font(size = "16")

class field_object:

    def __init__(self, original_field, cclf_field, cleaned_field, idr_table_name):
        self.original_field = original_field
        self.cclf_field = cclf_field
        self.cleaned_field = cleaned_field
        self.idr_table_name = idr_table_name
        self.idr_view_name = "V2_MDCR_" + idr_table_name
class mapping_object:
    def __init__(self):
        self.bfd_original_field = ""
        self.cclf_field = ""
        self.bfd_cleaned_field = ""
        self.idr_columns = set()
        self.idr_tables = set()
        self.idr_views = set()

    def set_bfd_original_field(self, bfd_label):
        self.bfd_original_field = bfd_label

    def get_bfd_original_field(self):
        return self.bfd_original_field

    def set_cclf_field(self, cclf_field):
        self.cclf_field = cclf_field

    def get_cclf_field(self):
        return self.cclf_field

    def set_cclf_field(self, cclf_field):
        self.cclf_field = cclf_field

    def get_cclf_field(self):
        return self.cclf_field

    def set_bfd_cleaned_field(self, bfd_label):
        self.bfd_cleaned_field = bfd_label

    def get_bfd_cleaned_field(self):
        return self.bfd_cleaned_field

    def add_idr_tables(self, idr_table):
        self.idr_tables.add(idr_table)

    def add_idr_columns(self, idr_column):
        self.idr_columns.add(idr_column)

    def get_idr_columns(self):
        return self.idr_columns

    def get_idr_tables(self):
        return self.idr_tables

    def add_idr_views(self, idr_view):
        self.idr_views.add(idr_view)

    def get_idr_views(self):
        return self.idr_views

# IDR
idr_wb = load_workbook('/Users/gkmf/bfd/idr-investigation/data-dictionaries/SF-IDR-Data-Dictionary-R204-2024-05-30.xlsx')
idr_sheet = idr_wb["Column_Defs"]
idr_column_names = idr_sheet['F'] # IDR Column Name
idr_fields_non_ss = []

ws = idr_wb.active
row = 5

for cell in idr_column_names:
    if cell.value is not None and "MCS" not in cell.value and "FISS" not in cell.value:
        idr_table_name = ws.cell(row=row, column=5).value #idr table name
        idr_entity_name = ws.cell(row=row, column=2).value # Entity Name
        if (idr_table_name is None):
            idr_table_name = ""
        if (idr_entity_name is None):
            idr_entity_name = ""
        if ("MCS" not in idr_table_name and "FISS" not in idr_table_name
                and "MDCD" not in idr_table_name and "medicaid" not in idr_entity_name.lower()):
            cleaned_field = cell.value.lower().replace("_", "").replace(".", "").replace("-", "").strip()
            idr_fields_non_ss.append(field_object(cell.value, "", cleaned_field, idr_table_name))
    row += 1

print("idr_fields_non_ss: " + str(len(idr_fields_non_ss))) # 21,323 with SS fields, 18625 w/o SS fields

# BFD
bfd_wb = load_workbook('/Users/gkmf/bfd/idr-investigation/data-dictionaries/BFD-Data-Dictionary.xlsx')
bfd_sheet = bfd_wb["V2"]
bfd_column_fields = bfd_sheet["O"] # Grab the CCW Mapping column
bfd_fields = []

ws = bfd_wb.active
row = 2

for cell in bfd_column_fields:
    if cell.value == "CCW Mapping":
        continue

    if (cell.value != None and cell.value != ""):
        cleaned_field = cell.value.replace("1ST", "1").replace("2ND", "2").replace("3RD", "3").replace("4TH", "4")
        cleaned_field = cleaned_field.lower().replace("_", "").replace(".", "").replace("-", "").strip()
        cclf_mapping_field = ws.cell(row=row, column=16).value #CCLF Mapping

        if (cclf_mapping_field != None and cclf_mapping_field != ""):
            cclf_mapping_field_split = cclf_mapping_field.split(".")
            cclf_mapping_field_cleaned = cclf_mapping_field_split[-1].lower().replace("_", "").replace(".", "").replace("-", "").strip()
            bfd_fields.append(field_object(cell.value, cclf_mapping_field, cclf_mapping_field_cleaned, ""))
        else:
            # No CCLF field associated
            bfd_fields.append(field_object(cell.value, "", cleaned_field, ""))
    row += 1

#print("bfd_fields: " + str(len(bfd_fields))) # 602

# Map BFD's CCW fields to IDR fields
mappings = []

for bfd_field_object in bfd_fields:
    mapping = mapping_object()
    mapping.set_bfd_original_field(bfd_field_object.original_field)
    mapping.set_cclf_field(bfd_field_object.cclf_field)
    mapping.set_bfd_cleaned_field(bfd_field_object.cleaned_field)

    for idr_field_object in idr_fields_non_ss:
        if ((bfd_field_object.cleaned_field == idr_field_object.cleaned_field)
                or bfd_field_object.cleaned_field in idr_field_object.cleaned_field):
            mapping.add_idr_columns(idr_field_object.original_field)
            mapping.add_idr_tables(idr_field_object.idr_table_name)
            mapping.add_idr_views(idr_field_object.idr_view_name)
        elif fuzz.token_set_ratio(bfd_field_object.cleaned_field, idr_field_object.cleaned_field) >= 70:
            mapping.add_idr_columns(idr_field_object.original_field)
            mapping.add_idr_tables(idr_field_object.idr_table_name)
            mapping.add_idr_views(idr_field_object.idr_view_name)

    mappings.append(mapping)

full_mapping_output_filename = "/Users/gkmf/bfd/idr-investigation/output/IDR_BFD_CCW_mapping.xlsx"
wb_full_mapping = load_workbook(filename=full_mapping_output_filename)
ws = wb_full_mapping.active

# Output CCW-IDR Mappings

# Format Headers
ws.cell(row=1, column=1, value="BFD's original CCW field").font = header_font_style
ws.cell(row=1, column=2, value="BFD's CCLF Mapping").font = header_font_style
ws.cell(row=1, column=3, value="BFD's cleaned CCW field").font = header_font_style
ws.cell(row=1, column=4, value="Possible IDR Column Name").font = header_font_style
ws.cell(row=1, column=5, value="Possible IDR Table Name").font = header_font_style
ws.cell(row=1, column=6, value="Possible IDR View Name").font = header_font_style

ws['A1'].alignment = Alignment(wrap_text=True)
ws['B1'].alignment = Alignment(wrap_text=True)
ws['C1'].alignment = Alignment(wrap_text=True)
ws['D1'].alignment = Alignment(wrap_text=True)
ws['E1'].alignment = Alignment(wrap_text=True)
ws['F1'].alignment = Alignment(wrap_text=True)

ws['A1'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type = "solid")
ws['B1'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type = "solid")
ws['C1'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type = "solid")
ws['D1'].fill = PatternFill(start_color="90EE90", end_color="FFC7CE", fill_type = "solid")
ws['E1'].fill = PatternFill(start_color="90EE90", end_color="FFC7CE", fill_type = "solid")
ws['F1'].fill = PatternFill(start_color="90EE90", end_color="FFC7CE", fill_type = "solid")

row = 2
for mapping in mappings:
    ws.cell(row=row, column=1, value=mapping.get_bfd_original_field())
    ws.cell(row=row, column=2, value=mapping.get_cclf_field())
    ws.cell(row=row, column=3, value=mapping.get_bfd_cleaned_field())
    ws['A' + str(row)].alignment = Alignment(wrap_text=True)
    ws['B' + str(row)].alignment = Alignment(wrap_text=True)
    ws['C' + str(row)].alignment = Alignment(wrap_text=True)

    # Add IDR Columns
    if (len(mapping.get_idr_columns()) == 0):
        ws.cell(row=row, column=4, value="")
    else:
        formatted_idr_columns = ""
        idr_columns = mapping.get_idr_columns()
        for idr_column in idr_columns:
            formatted_idr_columns += idr_column + ","
        ws.cell(row=row, column=4, value=formatted_idr_columns)
        ws['C' + str(row)].alignment = Alignment(wrap_text=True)

    # Add IDR Tables
    if (len(mapping.get_idr_tables()) == 0):
        ws.cell(row=row, column=5, value="")
    else:
        formatted_idr_tables = ""
        idr_tables = mapping.get_idr_tables()
        for idr_table in idr_tables:
            formatted_idr_tables += idr_table + ","
        ws.cell(row=row, column=5, value=formatted_idr_tables)
        ws['D' + str(row)].alignment = Alignment(wrap_text=True)

    # Add IDR Views
    if (len(mapping.get_idr_views()) == 0):
        ws.cell(row=row, column=6, value="")
    else:
        formatted_idr_views = ""
        idr_views = mapping.get_idr_views()
        for idr_view in idr_views:
            formatted_idr_views += idr_view + ","
        ws.cell(row=row, column=6, value=formatted_idr_views)
        ws['E' + str(row)].alignment = Alignment(wrap_text=True)

    row += 1

wb_full_mapping.close()
wb_full_mapping.save(filename=full_mapping_output_filename)

'''
print(fuzz.token_sort_ratio("AT_PHYSN_UPIN", "CLM_ATNDG_PRVDR_UPIN_NUM")) # ratio 49
print(fuzz.token_sort_ratio("Claim Attending Provider Unique Provider Identification Number Number", "Attending Physician UPIN")) # ratio 41
#attendingPhysicianUpin
print(fuzz.token_sort_ratio("AT_PHYSN_UPIN", "ATNDG_PRVDR_UPIN_NUM")) # ratio 55 if remove CLM
'''
