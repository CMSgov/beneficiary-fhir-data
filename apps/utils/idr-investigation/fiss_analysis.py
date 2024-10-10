from openpyxl import load_workbook

class field_object:
    def __init__(self, original_field, cleaned_field):
        self.original_field = original_field
        self.cleaned_field = cleaned_field

    def get_original_field(self):
        return self.original_field

    def get_cleaned_field(self):
        return self.cleaned_field
class mapping_object:
    def __init__(self):
        self.bfd_original_field = ""
        self.bfd_cleaned_field = ""
        self.rda_original_source_copy_field = ""
        self.rda_cleaned_source_copy_field = ""
        self.idr_original_fields = set()

    def set_bfd_original_field(self, bfd_label):
        self.bfd_original_field = bfd_label

    def get_bfd_original_field(self):
        return self.bfd_original_field

    def set_bfd_cleaned_field(self, bfd_label):
        self.bfd_cleaned_field = bfd_label

    def get_bfd_cleaned_field(self):
        return self.bfd_cleaned_field

    def set_rda_original_source_copy_field(self, rda_label):
        self.rda_original_source_copy_field = rda_label

    def get_rda_original_source_copy_field(self):
        return self.rda_original_source_copy_field

    def set_rda_cleaned_source_copy_field(self, rda_label):
        self.rda_cleaned_source_copy_field = rda_label

    def get_rda_cleaned_source_copy_field(self):
        return self.rda_cleaned_source_copy_field

    def add_idr_original_field(self, idr_label):
        self.idr_original_fields.add(idr_label)

    def get_idr_original_fields(self):
        return self.idr_original_fields

# RDA
rda_wb = load_workbook('/Users/gkmf/bfd/idr-investigation/data-dictionaries/RDA API Data Dictionary.xlsx')
rda_sheet = rda_wb["B) Data Dictionary"]
rda_column_fields = rda_sheet['C'] # Grab the column
rda_fields = []
for cell in rda_column_fields:
    cleaned_field = ''.join(e for e in cell.value if e.isalnum()).lower()
    rda_fields.append(field_object(cell.value, cleaned_field))

# IDR
idr_wb = load_workbook('/Users/gkmf/bfd/idr-investigation/data-dictionaries/SF-IDR-Data-Dictionary-R204-2024-05-30.xlsx')
idr_sheet = idr_wb["Column_Defs"]
idr_column_fields = idr_sheet['L']
idr_fields = []

for cell in idr_column_fields:
    if cell.value is not None and cell.value.lower() not in ['idr derived', 'derived', 'unknown', '-', ' ']:
        cleaned_field = cell.value.lower().replace("_", "").replace(".", "").replace("-", "").strip()
        idr_fields.append(field_object(cell.value, cleaned_field))

# BFD
bfd_wb = load_workbook('/Users/gkmf/bfd/idr-investigation/data-dictionaries/all_rda_columns.xlsx')
bfd_sheet = bfd_wb["Sheet1"]
bfd_column_fields = bfd_sheet['A'] # Grab the column
bfd_fields = []

for cell in bfd_column_fields:
    cleaned_field = ''.join(e for e in cell.value if e.isalnum()).lower()
    bfd_fields.append(field_object(cell.value, cleaned_field))

# Map BFD's rda fields to RDA's source_copybook_field_labels
bfd_to_rda_source_field_mapping = []
unique_fields = set()

for bfd_field_object in bfd_fields:
    mapping = mapping_object()

    for rda_field_object in rda_fields:
        if (bfd_field_object.cleaned_field not in unique_fields
                and bfd_field_object.cleaned_field in rda_field_object.cleaned_field):
            unique_fields.add(bfd_field_object.cleaned_field)
            mapping.set_bfd_original_field(bfd_field_object.original_field)
            mapping.set_bfd_cleaned_field(bfd_field_object.cleaned_field)
            mapping.set_rda_original_source_copy_field(rda_field_object.original_field)
            mapping.set_rda_cleaned_source_copy_field(rda_field_object.cleaned_field)

            bfd_to_rda_source_field_mapping.append(mapping)
    if mapping.get_bfd_cleaned_field() == "":
        mapping.set_bfd_original_field(bfd_field_object.original_field)
        mapping.set_bfd_cleaned_field(bfd_field_object.cleaned_field)
        bfd_to_rda_source_field_mapping.append(mapping)

output_filename = "/Users/gkmf/bfd/idr-investigation/output/2.0_BFD_RDA_matches.xlsx"
wb_output = load_workbook(filename=output_filename)
ws = wb_output.active

#headers
ws.cell(row=1, column=1, value="BFD's original RDA field")
ws.cell(row=1, column=2, value="BFD's cleaned RDA field")
ws.cell(row=1, column=3, value="RDA's original upstream source field")
ws.cell(row=1, column=4, value="RDA's cleaned upstream source field")

row = 2
for mapping in bfd_to_rda_source_field_mapping:
    ws.cell(row=row, column=1, value=mapping.get_bfd_original_field())
    ws.cell(row=row, column=2, value=mapping.get_bfd_cleaned_field())
    if mapping.get_rda_cleaned_source_copy_field() == "":
        ws.cell(row=row, column=3, value="")
    else:
        ws.cell(row=row, column=3, value=mapping.get_rda_original_source_copy_field())
        ws.cell(row=row, column=4, value=mapping.get_rda_cleaned_source_copy_field())

    row += 1

wb_output.close()
wb_output.save(filename=output_filename)

# Add manually identified fields to mapping object
input_filename = "/Users/gkmf/bfd/idr-investigation/output/BFD_RDA_matches (+ Manual Additions).xlsx"
wb_output = load_workbook(filename=input_filename)
ws = wb_output.active
rows = len(ws['A'])
mappings = []

for row in range(2, rows+1):
    bfd_original_field = ws.cell(row=row, column=1).value
    bfd_cleaned_field = ws.cell(row=row, column=2).value
    rda_original_source_copy_field = ws.cell(row=row, column=3).value
    rda_cleaned_source_copy_field = ws.cell(row=row, column=4).value
    manually_identified_field = ws.cell(row=row, column=5).value

    mapping = mapping_object()
    mapping.set_bfd_original_field(bfd_original_field)
    mapping.set_bfd_cleaned_field(bfd_cleaned_field)

    if manually_identified_field is not None and manually_identified_field != "System Derived" and manually_identified_field != "Missing?":
        cleaned_manual_field = ''.join(e for e in manually_identified_field if e.isalnum()).lower()
        mapping.set_rda_original_source_copy_field(cleaned_manual_field)
    else:
        mapping.set_rda_original_source_copy_field(rda_original_source_copy_field)
        mapping.set_rda_cleaned_source_copy_field(rda_cleaned_source_copy_field)

    mappings.append(mapping)

# Map RDA's source_copybook_field_labels to IDR's fields

for mapping in mappings:
    for idr_field_object in idr_fields:
        rda_cleaned_source_copy_field = mapping.get_rda_cleaned_source_copy_field()
        if (rda_cleaned_source_copy_field is not None
                and rda_cleaned_source_copy_field != ""
                and rda_cleaned_source_copy_field != " "
                and rda_cleaned_source_copy_field.replace('idr', '') in idr_field_object.cleaned_field):
            mapping.add_idr_original_field(idr_field_object.original_field)

full_mapping_output_filename = "/Users/gkmf/bfd/idr-investigation/output/full_mapping.xlsx"
wb_full_mapping = load_workbook(filename=full_mapping_output_filename)
ws = wb_full_mapping.active

#headers
ws.cell(row=1, column=1, value="BFD's original RDA field")
ws.cell(row=1, column=2, value="BFD's cleaned RDA field")
ws.cell(row=1, column=3, value="RDA's original upstream source field")
ws.cell(row=1, column=4, value="RDA's cleaned upstream source field")
ws.cell(row=1, column=5, value="IDR's field")

row = 2
for mapping in mappings:
    ws.cell(row=row, column=1, value=mapping.get_bfd_original_field())
    ws.cell(row=row, column=2, value=mapping.get_bfd_cleaned_field())
    ws.cell(row=row, column=3, value=mapping.get_rda_original_source_copy_field())
    ws.cell(row=row, column=4, value=mapping.get_rda_cleaned_source_copy_field())

    if (len(mapping.get_idr_original_fields()) == 0):
        ws.cell(row=row, column=5, value="")
    else:
        formatted_matched_idr_fields = ""
        idr_fields = mapping.get_idr_original_fields()
        for idr_field in idr_fields:
            formatted_matched_idr_fields += idr_field + ","
        ws.cell(row=row, column=5, value=formatted_matched_idr_fields)

    row += 1

wb_full_mapping.close()
wb_full_mapping.save(filename=full_mapping_output_filename)
